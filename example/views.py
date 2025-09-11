from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from .forms import StudentForm, InvoiceSettingsForm
from django.contrib.auth.models import User
from django.db.models import Count, Sum, F, Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import transaction
from .models import CSRProfile, Course, Batch, Student, InvoiceSettings,StudentInvoice
from .utils import render_printable_invoice
import json
import io
import csv
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def calculate_date_range_revenue(students, start_date=None, end_date=None):
    """Helper function to calculate revenue based on payments received within date range"""
    batch_groups = {}
    course_groups = {}
    
    for student in students:
        # Calculate payments within date range for this student
        advance_in_range = 0
        second_installment_in_range = 0
        
        if start_date and end_date:
            if isinstance(start_date, str):
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                start_date_obj = start_date
                end_date_obj = end_date
            
            if student.created_at and start_date_obj <= student.created_at.date() <= end_date_obj:
                advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            if student.due_date and start_date_obj <= student.due_date <= end_date_obj:
                second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        elif start_date:
            if isinstance(start_date, str):
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            else:
                start_date_obj = start_date
            
            if student.created_at and student.created_at.date() >= start_date_obj:
                advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            if student.due_date and student.due_date >= start_date_obj:
                second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        elif end_date:
            if isinstance(end_date, str):
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                end_date_obj = end_date
            
            if student.created_at and student.created_at.date() <= end_date_obj:
                advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            if student.due_date and student.due_date <= end_date_obj:
                second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        else:
            # No date filter, include all payments
            advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        
        payment_in_range = advance_in_range + second_installment_in_range
        
        # Group by batch
        batch_name = student.batch.batch_number if student.batch else 'N/A'
        if batch_name not in batch_groups:
            batch_groups[batch_name] = {
                'total_revenue': 0,
                'received_payment': 0,
                'pending_payment': 0,
                'student_count': 0
            }
        
        batch_groups[batch_name]['total_revenue'] += float(student.discounted_price) if student.discounted_price else 0
        batch_groups[batch_name]['received_payment'] += payment_in_range
        batch_groups[batch_name]['pending_payment'] += float(student.balance) if student.balance else 0
        batch_groups[batch_name]['student_count'] += 1
        
        # Group by course
        for course in student.courses.all():
            course_name = course.name
            if course_name not in course_groups:
                course_groups[course_name] = {
                    'total_revenue': 0,
                    'received_payment': 0,
                    'pending_payment': 0,
                    'student_count': 0
                }
            
            course_groups[course_name]['total_revenue'] += float(student.discounted_price) if student.discounted_price else 0
            course_groups[course_name]['received_payment'] += payment_in_range
            course_groups[course_name]['pending_payment'] += float(student.balance) if student.balance else 0
            course_groups[course_name]['student_count'] += 1
    
    # Convert to list format for template compatibility
    batch_revenue = [{'batch__batch_number': k, **v} for k, v in batch_groups.items()]
    course_revenue = [{'courses__name': k, **v} for k, v in course_groups.items()]
    
    return batch_revenue, course_revenue


# Custom JSON encoder to handle Decimal objects
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super(DecimalEncoder, self).default(obj)

# Create your views here.

def login_view(request):
    """Custom login view with cool design"""
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        else:
            # Check if user is a CSR
            try:
                csr = request.user.csr_profile
                return redirect('csr_dashboard')
            except:
                messages.info(request, 'You are logged in but not assigned to any role.')
                return redirect('login')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome, {username}!')
            
            if user.is_staff:
                return redirect('admin_dashboard')
            else:
                # Check if user is a CSR
                try:
                    csr = user.csr_profile
                    return redirect('csr_dashboard')
                except:
                    messages.warning(request, 'Your account is not assigned to any role. Please contact an administrator.')
                    return redirect('login')
        else:
            messages.error(request, 'Username or password is incorrect.')
    
    return render(request, 'login.html')

def logout_view(request):
    """Custom logout view"""
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('login')

# Toggle dark mode
def toggle_dark_mode(request):
    """Toggle dark mode in session"""
    if request.method == 'POST':
        if request.session.get('dark_mode', False):
            request.session['dark_mode'] = False
        else:
            request.session['dark_mode'] = True
        return JsonResponse({'status': 'success', 'dark_mode': request.session['dark_mode']})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# Admin Dashboard Views
@staff_member_required(login_url='login')
def admin_dashboard(request):
    """Admin dashboard view"""
    # Get counts for dashboard stats
    total_csrs = CSRProfile.objects.count()
    total_courses = Course.objects.count()
    total_students = Student.objects.count()  # All students (active + inactive batches)
    current_students = Student.objects.filter(batch__status='active').count()  # Students in active batches only
    total_batches = Batch.objects.count()
    active_batches = Batch.objects.filter(status='active').count()
    inactive_batches = Batch.objects.filter(status='inactive').count()
    
    # Total Revenue (sum of all enrolled-course fees, using discounted price)
    total_revenue = Student.objects.all().aggregate(
        total=Sum('discounted_price')
    )['total'] or 0
    
    # Payment Received (sum of total_amount - actual payments received)
    payment_received = Student.objects.all().aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Pending Payments (sum of balance - remaining amounts to be paid)
    pending_payments = Student.objects.all().aggregate(total=Sum('balance'))['total'] or 0
    
    # Students per Batch (info-cards per batch with student count)
    batches_with_students = Batch.objects.annotate(
        student_count=Count('students')
    ).order_by('-created_at')
    
    # Get recent students for activity feed
    recent_students = Student.objects.select_related('created_by', 'batch').order_by('-created_at')[:5]
    
    # Get courses with student counts for course distribution chart
    courses_with_counts = Course.objects.annotate(
        student_count=Count('students')
    ).order_by('-student_count')
    
    # Prepare course data for chart.js
    course_labels = [course.name for course in courses_with_counts]
    course_data = [course.student_count for course in courses_with_counts]
    course_colors = [
        f'rgba({200 + i}, {200 + i}, {200 + i}, {min(0.9 - (i * 0.1), 0.9)})'
        for i in range(len(courses_with_counts))
    ]
    courses_data = {
        'labels': course_labels,
        'data': course_data,
        'colors': course_colors
    }
    
    # Get CSR performance data (number of students enrolled by each CSR)
    csr_performance = CSRProfile.objects.annotate(
        students_enrolled=Count('students')
    ).order_by('-students_enrolled')[:5]
    
    # Prepare CSR performance data for chart.js
    csr_labels = [csr.full_name for csr in csr_performance]
    csr_data = [csr.students_enrolled for csr in csr_performance]
    csr_performance_data = {
        'labels': csr_labels,
        'data': csr_data
    }
    
    # Get monthly revenue data for the chart
    current_year = timezone.now().year
    monthly_revenue = []
    
    for month in range(1, 13):
        month_revenue = Student.objects.filter(
            created_at__year=current_year,
            created_at__month=month
        ).aggregate(total=Sum('discounted_price'))['total'] or 0
        monthly_revenue.append(month_revenue)
    
    context = {
        'total_csrs': total_csrs,
        'total_courses': total_courses,
        'total_students': total_students,
        'current_students': current_students,
        'total_batches': total_batches,
        'active_batches': active_batches,
        'inactive_batches': inactive_batches,
        'total_revenue': total_revenue,
        'payment_received': payment_received,
        'pending_payments': pending_payments,
        'batches_with_students': batches_with_students,
        'recent_students': recent_students,
        'courses_with_counts': courses_with_counts,
        'csr_performance': csr_performance,
        'monthly_revenue': json.dumps(monthly_revenue, cls=DecimalEncoder),
        'courses_data': json.dumps(courses_data, cls=DecimalEncoder),
        'csr_performance_data': json.dumps(csr_performance_data, cls=DecimalEncoder),
    }
    
    return render(request, 'invoice/admin_dashboard.html', context)

@staff_member_required(login_url='login')
def csr_management(request):
    """CSR management view"""
    # Get all CSRs
    csrs = CSRProfile.objects.all().order_by('-date_joined')
    
    # Count active and inactive CSRs
    active_csrs = csrs.filter(is_active=True).count()
    inactive_csrs = csrs.filter(is_active=False).count()
    
    # Get top performing CSRs (by invoice count)
    # In a real application, we would sort by actual invoice count
    # For now, we'll just use the first 5 CSRs as an example
    top_csrs = csrs.filter(is_active=True)[:5]
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            # Handle CSR creation
            username = request.POST.get('username')
            password = request.POST.get('password')
            full_name = request.POST.get('full_name')
            is_active = request.POST.get('is_active') == 'on'
            lead_role = request.POST.get('lead_role') == 'on'
            
            # Check if username already exists
            if User.objects.filter(username=username).exists():
                messages.error(request, f'Username {username} already exists.')
                return redirect('csr_management')
            
            try:
                with transaction.atomic():
                    # Create user
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        is_staff=False,  # Not a staff user (admin)
                        is_active=is_active  # Can be active or inactive based on admin selection
                    )
                    
                    # Create CSR profile
                    csr = CSRProfile.objects.create(
                        user=user,
                        full_name=full_name,
                        is_active=is_active,
                        lead_role=lead_role
                    )
                    
                    messages.success(request, f'CSR {full_name} created successfully. They can now log in to access the CSR dashboard.')
            except Exception as e:
                messages.error(request, f'Error creating CSR: {str(e)}')
        
        elif action == 'edit':
            # Handle CSR edit
            csr_id = request.POST.get('csr_id')
            full_name = request.POST.get('full_name')
            username = request.POST.get('username')
            password = request.POST.get('password')
            is_active = request.POST.get('is_active') == 'on'
            lead_role = request.POST.get('lead_role') == 'on'
            
            try:
                with transaction.atomic():
                    # Get CSR profile
                    csr = get_object_or_404(CSRProfile, id=csr_id)
                    
                    # Update CSR profile
                    csr.full_name = full_name
                    csr.is_active = is_active
                    csr.lead_role = lead_role
                    csr.save()
                    
                    # Update user
                    user = csr.user
                    user.username = username
                    if password:  # Only update password if provided
                        user.set_password(password)
                    user.is_active = is_active
                    user.save()
                    
                    messages.success(request, f'CSR {full_name} updated successfully.')
            except Exception as e:
                messages.error(request, f'Error updating CSR: {str(e)}')
        
        elif action == 'delete':
            # Handle CSR delete
            csr_id = request.POST.get('csr_id')
            
            try:
                # Get CSR profile
                csr = CSRProfile.objects.get(id=csr_id)
                user = csr.user
                full_name = csr.full_name
                
                # Delete user (will cascade to CSR profile)
                user.delete()
                
                messages.success(request, f'CSR {full_name} deleted successfully.')
            except CSRProfile.DoesNotExist:
                messages.error(request, 'CSR not found.')
            except Exception as e:
                messages.error(request, f'Error deleting CSR: {str(e)}')
        
        return redirect('csr_management')
    
    context = {
        'csrs': csrs,
        'active_csrs': active_csrs,
        'inactive_csrs': inactive_csrs,
        'top_csrs': top_csrs,
    }
    
    return render(request, 'invoice/csr_management.html', context)

@staff_member_required(login_url='login')
def course_management(request):
    """Course management view for admin users"""
    # Check if user is an admin
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access course management.')
        return redirect('dashboard')
        
    # Get all courses
    courses = Course.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            # Handle course creation
            name = request.POST.get('name')
            trainer_name = request.POST.get('trainer_name')
            price = request.POST.get('price')
            durations = request.POST.getlist('duration')
            duration = ','.join(durations) if durations else ''
            
            try:
                # Create course
                course = Course.objects.create(
                    name=name,
                    trainer_name=trainer_name,
                    price=price,
                    duration=duration
                )
                
                messages.success(request, f'Course {name} created successfully.')
            except Exception as e:
                messages.error(request, f'Error creating course: {str(e)}')
        
        elif action == 'edit':
            # Handle course edit
            course_id = request.POST.get('course_id')
            name = request.POST.get('name')
            trainer_name = request.POST.get('trainer_name')
            price = request.POST.get('price')
            durations = request.POST.getlist('edit_duration')
            duration = ','.join(durations) if durations else ''  # Note: using edit_duration from form
            
            try:
                course = Course.objects.get(id=course_id)
                course.name = name
                course.trainer_name = trainer_name
                course.price = price
                course.duration = duration
                course.save()
                
                messages.success(request, f'Course {name} updated successfully.')
            except Course.DoesNotExist:
                messages.error(request, 'Course not found.')
            except Exception as e:
                messages.error(request, f'Error updating course: {str(e)}')
        
        elif action == 'delete':
            # Handle course delete
            course_id = request.POST.get('course_id')
            
            try:
                course = Course.objects.get(id=course_id)
                name = course.name
                course.delete()
                
                messages.success(request, f'Course {name} deleted successfully.')
            except Course.DoesNotExist:
                messages.error(request, 'Course not found.')
            except Exception as e:
                messages.error(request, f'Error deleting course: {str(e)}')
        
        return redirect('course_management')
    
    context = {
        'courses': courses,
        'csr': None,
    }
    
    return render(request, 'invoice/course_management.html', context)


def csr_course_management(request):
    """Course management view for CSR users with lead role"""
    # Check if user is a CSR with lead role
    try:
        # Get CSR profile
        csr = CSRProfile.objects.get(user=request.user)
        if not csr.lead_role:
            messages.error(request, 'You do not have permission to access course management.')
            return redirect('csr_dashboard')
    except CSRProfile.DoesNotExist:
        messages.error(request, 'You do not have permission to access course management.')
        return redirect('csr_dashboard')
        
    # Get all courses
    courses = Course.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            # Handle course creation
            name = request.POST.get('name')
            trainer_name = request.POST.get('trainer_name')
            price = request.POST.get('price')
            durations = request.POST.getlist('duration')
            duration = ','.join(durations) if durations else ''
            
            try:
                # Create course
                course = Course.objects.create(
                    name=name,
                    trainer_name=trainer_name,
                    price=price,
                    duration=duration
                )
                
                messages.success(request, f'Course {name} created successfully.')
            except Exception as e:
                messages.error(request, f'Error creating course: {str(e)}')
        
        elif action == 'edit':
            # Handle course edit
            course_id = request.POST.get('course_id')
            name = request.POST.get('name')
            trainer_name = request.POST.get('trainer_name')
            price = request.POST.get('price')
            durations = request.POST.getlist('edit_duration')
            duration = ','.join(durations) if durations else ''  # Note: using edit_duration from form
            
            try:
                course = Course.objects.get(id=course_id)
                course.name = name
                course.trainer_name = trainer_name
                course.price = price
                course.duration = duration
                course.save()
                
                messages.success(request, f'Course {name} updated successfully.')
            except Course.DoesNotExist:
                messages.error(request, 'Course not found.')
            except Exception as e:
                messages.error(request, f'Error updating course: {str(e)}')
        
        elif action == 'delete':
            # Handle course delete
            course_id = request.POST.get('course_id')
            
            try:
                course = Course.objects.get(id=course_id)
                name = course.name
                course.delete()
                
                messages.success(request, f'Course {name} deleted successfully.')
            except Course.DoesNotExist:
                messages.error(request, 'Course not found.')
            except Exception as e:
                messages.error(request, f'Error deleting course: {str(e)}')
        
        return redirect('course_management_csr')
    
    context = {
        'courses': courses,
        'csr': csr,
    }
    
    return render(request, 'invoice/course_management.html', context)

@staff_member_required(login_url='login')
def admin_settings(request):
    """Admin settings view"""
    return render(request, 'invoice/admin_settings.html')

# CSR Dashboard Views
@login_required(login_url='login')
def csr_dashboard(request):
    """CSR dashboard view"""
    # Check if user is a CSR
    try:
        # Get the CSR profile using the related_name 'csr_profile'
        csr = CSRProfile.objects.get(user=request.user)
    except CSRProfile.DoesNotExist:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('login')
    print('csr',csr.lead_role)
    # Get counts for dashboard stats
    total_batches = Batch.objects.all().count()  # Show all batches count
    active_batches = Batch.objects.filter(status='active').count()  # Active batches count
    
    # Get CSR count for lead CSRs (same as admin dashboard)
    total_csrs = CSRProfile.objects.count() if (csr.lead_role or request.user.is_superuser) else None
    
    # For students, show all for lead/admin, both total and own for regular CSRs
    if csr.lead_role or request.user.is_superuser:
        total_students = Student.objects.count()
        csr_students = None  # Not needed for lead CSRs
        current_students = Student.objects.filter(batch__status='active').count()  # Students in active batches
        
        # Add revenue and pending payment data for lead CSRs
        # Total Revenue (sum of all enrolled-course fees, using discounted price)
        total_revenue = Student.objects.all().aggregate(
            total=Sum('discounted_price')
        )['total'] or 0
        
        # Payment Received (sum of total_amount - actual payments received)
        payment_received = Student.objects.all().aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Pending Payments (sum of balance - remaining amounts to be paid)
        pending_payments = Student.objects.all().aggregate(total=Sum('balance'))['total'] or 0
    else:
        # For regular CSRs, show both total students and their own students
        total_students = Student.objects.all().count()
        csr_students = Student.objects.filter(created_by=csr).count()
        current_students = Student.objects.filter(created_by=csr, batch__status='active').count()
        # For regular CSRs, we don't calculate revenue metrics
        total_revenue = None
        payment_received = None
        pending_payments = None
        
    total_courses = Course.objects.count()
    
    # Get recent students for display in the dashboard
    recent_students = Student.objects.filter(created_by=csr).order_by('-created_at')[:5]
    
    # Get courses for lead role CSRs
    courses = []
    if csr.lead_role or request.user.is_superuser:
        courses = Course.objects.all().order_by('-created_at')[:5]
    
    context = {
        'csr': csr,
        'total_batches': total_batches,
        'active_batches': active_batches,
        'total_students': total_students,
        'current_students': current_students,
        'csr_students': csr_students,  # Add CSR's own student count for regular CSRs
        'total_courses': total_courses,
        'total_csrs': total_csrs,  # Add total CSRs count for lead CSRs
        'students': recent_students,  # Add recent students to context
        'courses': courses,  # Add courses for lead role CSRs
    }
    
    # Add revenue data to context for lead CSRs
    print(f"DEBUG - Lead role: {csr.lead_role}, Is superuser: {request.user.is_superuser}")
    print(f"DEBUG - Before update, context keys: {context.keys()}")
    
    if csr.lead_role or request.user.is_superuser:
        context.update({
            'total_revenue': total_revenue,
            'payment_received': payment_received,
            'pending_payments': pending_payments,
            'is_lead': True,  # Add explicit flag for template
        })
        print(f"DEBUG - After update, context keys: {context.keys()}")
        print(f"DEBUG - Revenue values: {total_revenue}, {payment_received}, {pending_payments}")
    else:
        print("DEBUG - User is not lead CSR or admin, not adding revenue data")
    
    return render(request, 'invoice/csr_dashboard.html', context)

@login_required(login_url='login')
def csr_batch_management(request):
    """CSR batch management view for lead CSRs"""
    # Check if user is a CSR with lead role
    try:
        csr = CSRProfile.objects.get(user=request.user)
        if not csr.lead_role:
            messages.error(request, 'You do not have permission to access batch management.')
            return redirect('csr_dashboard')
    except CSRProfile.DoesNotExist:
        messages.error(request, 'You do not have permission to access batch management.')
        return redirect('csr_dashboard')
    
    # Handle batch creation
    if request.method == 'POST':
        batch_number = request.POST.get('batch_number')
        
        if not batch_number:
            messages.error(request, 'Batch number is required.')
        elif Batch.objects.filter(batch_number=batch_number).exists():
            messages.error(request, 'A batch with this number already exists.')
        else:
            batch = Batch.objects.create(
                batch_number=batch_number,
                created_by=csr
            )
            messages.success(request, f'Batch {batch_number} created successfully.')
    
    # Get all batches with counts
    batches = Batch.objects.all().order_by('-created_at')
    total_batches = batches.count()
    active_batches = batches.filter(status='active').count()
    
    # Get student counts
    total_students = Student.objects.count()
    current_students = Student.objects.filter(batch__status='active').count()
    
    context = {
        'batches': batches,
        'total_batches': total_batches,
        'active_batches': active_batches,
        'total_students': total_students,
        'current_students': current_students,
        'csr': csr,
    }
    
    return render(request, 'invoice/csr_batch_management.html', context)

@login_required(login_url='login')
def batch_management(request):
    """Batch management view for CSRs"""
    # Check if user is a CSR
    try:
        csr = request.user.csr_profile
    except:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('login')
    
    # Handle batch creation
    if request.method == 'POST':
        batch_number = request.POST.get('batch_number')
        
        if not batch_number:
            messages.error(request, 'Batch number is required.')
        elif Batch.objects.filter(batch_number=batch_number).exists():
            messages.error(request, 'A batch with this number already exists.')
        else:
            batch = Batch.objects.create(
                batch_number=batch_number,
                created_by=csr
            )
            messages.success(request, f'Batch {batch_number} created successfully.')
    
    # Get all batches (not just those created by this CSR)
    batches = Batch.objects.all().order_by('-created_at')
    
    context = {
        'batches': batches,
    }
    
    return render(request, 'invoice/batch_management.html', context)

@login_required(login_url='login')
def student_management(request):
    """Student management view for CSRs"""
    # Check if user is an admin or a CSR
    if request.user.is_superuser:
        # For admin users, show all batches
        batches = Batch.objects.all().order_by('-created_at')
        csr = None  # Admin doesn't have a CSR profile
    else:
        # For CSR users
        try:
            # Get CSR profile directly from the database
            csr = CSRProfile.objects.get(user=request.user)
            # Show all batches to all CSRs, regardless of role
            batches = Batch.objects.all().order_by('-created_at')
        except CSRProfile.DoesNotExist:
            messages.error(request, 'You are not authorized to access this page.')
            return redirect('login')
    courses = Course.objects.all().order_by('name')
    
    # Handle student creation
    if request.method == 'POST':
        name = request.POST.get('name')
        guardian_name = request.POST.get('guardian_name')
        phone_number = request.POST.get('phone_number')
        cnic = request.POST.get('cnic')
        batch_id = request.POST.get('batch')
        course_ids = request.POST.getlist('courses')
        discount_percent = request.POST.get('discount', 0)
        advance_payment = request.POST.get('advance_payment', 0)
        second_installment_due_date = request.POST.get('second_installment_due_date')
        schedule_choice = request.POST.get('schedule', 'weekend')
        
        # Validate required fields
        if not all([name, guardian_name, phone_number, cnic, batch_id, course_ids]):
            messages.error(request, 'Please fill in all required fields.')
        else:
            try:
                # Allow selection of any batch, not just those created by the current CSR
                batch = Batch.objects.get(id=batch_id)
                
                # Convert string values to appropriate types for Student model fields
                discount_percent = Decimal(discount_percent) if discount_percent else Decimal('0')
                advance_payment = int(float(advance_payment)) if advance_payment else 0
                
                # Calculate total fees from selected courses
                course_total = 0
                for course_id in course_ids:
                    try:
                        course = Course.objects.get(id=course_id)
                        course_total += course.price
                    except Course.DoesNotExist:
                        pass
                
                # Apply discount as percentage
                discount_amount = (course_total * discount_percent / Decimal('100'))
                discounted_price = course_total - discount_amount
                
                # Calculate second installment and balance
                second_installment = max(Decimal('0'), discounted_price - Decimal(str(advance_payment)))
                balance = second_installment  # Set balance equal to second_installment
                
                # Convert Decimal values to integers for Student model
                course_total_int = int(course_total)
                discounted_price_int = int(discounted_price)
                second_installment_int = int(second_installment)
                balance_int = int(balance)
                
                # Create student
                with transaction.atomic():
                    # Check if advance payment equals discounted price (fully paid)
                    # If so, set due_date equal to created_at since there are no dues left
                    due_date = None
                    if Decimal(str(advance_payment)) >= discounted_price:
                        due_date = timezone.now().date()
                    
                    # Set total_amount based on payment status
                    # If advance payment covers the full amount, consider it paid
                    payment_status = 'pending'
                    total_amount = advance_payment
                    
                    if Decimal(str(advance_payment)) >= discounted_price:
                        payment_status = 'paid'
                        total_amount = discounted_price  # Full amount is paid
                        
                    student = Student.objects.create(
                        name=name,
                        guardian_name=guardian_name,
                        phone_number=phone_number,
                        cnic=cnic,
                        batch=batch,
                        discount=discount_percent,
                        total_fees=course_total_int,
                        discounted_price=discounted_price_int,
                        advance_payment=advance_payment,
                        second_installment=second_installment_int,
                        balance=balance_int,  # Add balance field
                        total_amount=int(total_amount),  # Set total_amount correctly
                        payment_status=payment_status,  # Set payment status based on payment
                        second_installment_due_date=second_installment_due_date if second_installment_due_date else None,
                        due_date=due_date,  # Set due_date if fully paid
                        schedule=schedule_choice,
                        created_by=csr
                    )
                    
                    # Add courses
                    for course_id in course_ids:
                        try:
                            course = Course.objects.get(id=course_id)
                            student.courses.add(course)
                        except Course.DoesNotExist:
                            pass
                    
                    messages.success(request, f'Student {name} registered successfully.')
                    return redirect('student_management')
            except Batch.DoesNotExist:
                messages.error(request, 'Selected batch does not exist.')
            except ValueError:
                messages.error(request, 'Invalid numeric values provided.')
    
    # Get batch filter parameter if provided
    batch_filter = request.GET.get('batch')
    
    # Get students based on user role
    # For admin users and lead CSRs, show all students
    # For regular CSRs, only show their own students
    if request.user.is_superuser or (csr and csr.lead_role):
        students_query = Student.objects.all().select_related('batch')
    else:
        students_query = Student.objects.filter(created_by=csr).select_related('batch')
        
    # Apply batch filter if specified
    if batch_filter:
        students_query = students_query.filter(batch_id=batch_filter)
    
    students = students_query.order_by('-created_at')
    
    context = {
        'batches': batches,
        'courses': courses,
        'students': students,
        'csr': csr,  # Pass csr to template for sidebar visibility
    }
    
    return render(request, 'invoice/student_management.html', context)

@login_required(login_url='login')
def generate_invoice(request, student_id):
    """Generate printable invoice for a student"""
    # Get CSR profile
    try:
        # Get CSR profile directly from the database
        csr_profile = CSRProfile.objects.get(user=request.user)
    except CSRProfile.DoesNotExist:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('login')
    
    # Get student object
    student = get_object_or_404(Student, id=student_id, created_by=csr_profile)
    
    # Check if this is a pending invoice request
    is_pending = request.GET.get('pending', 'false').lower() == 'true'
    
    # For pending invoices, use today's date if payment is made before due date
    if is_pending and student.due_date and student.due_date > datetime.now().date():
        student.due_date = datetime.now().date()
    
    # Get or create StudentInvoice record for this student
    student_invoice, created = StudentInvoice.objects.get_or_create(student=student)
    
    # Determine which invoice settings will be used (same logic as get_invoice_context)
    invoice_settings = None

    # 1) Try student's creator invoice settings
    if student.created_by:
        invoice_settings = InvoiceSettings.objects.filter(csr=student.created_by).first()

    # 2) If not found, fall back to a Lead CSR with settings
    if not invoice_settings:
        invoice_settings = InvoiceSettings.objects.filter(csr__lead_role=True).first()

    # 3) As a final fallback, use (or create) settings for the logged-in CSR
    if not invoice_settings:
        invoice_settings, _ = InvoiceSettings.objects.get_or_create(
            csr=csr_profile,
            defaults={
                'current_serial_number': 1000
            }
        )

    # Only increment and persist the serial number if this is the first time generating this invoice type
    if not is_pending and student_invoice.present_invoice_no == 0:
        # Increment the serial number for regular invoice
        invoice_settings.current_serial_number += 1
        invoice_settings.save()
        
        # Save the invoice number to the student's invoice record
        student_invoice.present_invoice_no = invoice_settings.current_serial_number
        student_invoice.save()
    elif is_pending and student_invoice.pending_invoice_no == 0:
        # Increment the serial number for pending invoice
        invoice_settings.current_serial_number += 1
        invoice_settings.save()
        
        # Save the invoice number to the student's invoice record
        student_invoice.pending_invoice_no = invoice_settings.current_serial_number
        student_invoice.save()
    
    # Use the utility function to render the invoice
    return render_printable_invoice(request, student, is_pending=is_pending)

@login_required(login_url='login')
def edit_student(request, student_id):
    """Edit an existing student (CSR only) using ModelForm"""
    try:
        csr = request.user.csr_profile
    except:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('login')

    student = get_object_or_404(Student, id=student_id, created_by=csr)
    
    # Show all batches so CSR can reassign student to any existing batch
    batches = Batch.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        # Debug: Print POST data
        print(f"POST data: {request.POST}")
        print(f"POST courses: {request.POST.getlist('courses')}")
        
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            # Ensure the batch belongs to this CSR

            
            # Debug: Print form cleaned data
            print(f"Form cleaned data: {form.cleaned_data}")
            print(f"Form courses: {form.cleaned_data.get('courses')}")
            
            # Save instance without committing to handle M2M separately
            updated_student = form.save(commit=False)
            updated_student.save()
            form.save_m2m()
            print("Student updated and m2m saved. Schedule:", updated_student.schedule)
            print("Courses after save:", list(updated_student.courses.values_list('id', flat=True)))
            messages.success(request, 'Student updated successfully.')
            return redirect('student_management')
        else:
            # Debug: Print form errors
            print(f"Form errors: {form.errors}")
            # Return form errors as JSON for AJAX handling
            return JsonResponse({'errors': form.errors}, status=400)
    else:
        # For GET requests, return form data as JSON
        form = StudentForm(instance=student)
        
        # Get current courses for the student
        current_courses = list(student.courses.values_list('id', flat=True))
        
        # Prepare form data for JSON response
        form_data = {
            'id': student.id,
            'name': student.name,
            'guardian_name': student.guardian_name,
            'phone_number': student.phone_number,
            'cnic': student.cnic,
            'batch': student.batch.id,
            'courses': current_courses,
            'discount': int(student.discount),
            'total_fees': int(student.total_fees),
            'discounted_price': int(student.discounted_price),
            'advance_payment': int(student.advance_payment),
            'second_installment': int(student.second_installment),
            'balance': int(student.balance),  # Add balance field
            'total_amount': int(student.total_amount),  # Add total_amount field
            'payment_status': student.payment_status,  # Add payment_status field
            'second_installment_due_date': student.second_installment_due_date.isoformat() if student.second_installment_due_date else '',
            'due_date': student.due_date.isoformat() if student.due_date else '',
            'schedule': student.schedule,  # Add schedule field
            'payment_method': student.payment_method,  # Add payment method field
        }
        
        return JsonResponse({
            'student': form_data,
            'batches': list(batches.values('id', 'batch_number')),
            'courses': list(Course.objects.all().values('id', 'name', 'price')),
        })

@login_required(login_url='login')
def delete_student(request, student_id):
    """Delete a student (restricted). Only Superusers or Lead CSRs may delete."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    # Only Django superusers (admin) may delete
    if request.user.is_superuser:
        student = get_object_or_404(Student, id=student_id)
        student.delete()
        messages.success(request, 'Student deleted successfully.')
    else:
        messages.error(request, 'You do not have permission to delete students.')
    return redirect('student_management')

@login_required(login_url='login')
def update_payment_status(request, student_id):
    """Update payment status for a student (AJAX endpoint)"""
    # Get CSR profile
    try:
        csr_profile = request.user.csr_profile
    except:
        return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)
    
    # Get student object
    try:
        student = Student.objects.get(id=student_id, created_by=csr_profile)
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    
    # Process the request
    if request.method == 'POST':
        try:
            # Parse JSON data
            data = json.loads(request.body)
            payment_status = data.get('payment_status')
            
            # Validate payment status
            if payment_status not in ['paid', 'pending']:
                return JsonResponse({'success': False, 'error': 'Invalid payment status'}, status=400)
            
            # Update student payment status
            student.payment_status = payment_status
            
            # The Student.save() method will handle setting balance to 0 when payment_status is 'paid'
            # No need to manually set balance here as it's handled in the model's save method
                
            student.save()
            
            # Return the updated balance in the response
            return JsonResponse({
                'success': True,
                'balance': int(student.balance),
                'payment_status': student.payment_status
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

@login_required(login_url='login')
def invoice_settings(request):
    """Manage invoice settings (Lead CSR only)"""
    try:
        csr = request.user.csr_profile
    except:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('login')
    
    # Check if the CSR has lead role or is a superuser
    if not (csr.lead_role or request.user.is_superuser):
        messages.error(request, 'Only Lead CSRs can edit invoice settings.')
        return redirect('csr_dashboard')
    
    # Get or create invoice settings for this CSR
    # Using model defaults instead of hardcoded values
    settings, created = InvoiceSettings.objects.get_or_create(csr=csr)
    
    if request.method == 'POST':
        form = InvoiceSettingsForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'Invoice settings updated successfully.')
            return redirect('invoice_settings')
    else:
        form = InvoiceSettingsForm(instance=settings)
    
    context = {
        'form': form,
        'settings': settings,
        'csr': csr,
    }
    return render(request, 'invoice/invoice_settings.html', context)

# Batch Stats API Endpoint
def get_batch_stats(request):
    """API endpoint to get batch-wise student count and revenue data"""
    # Get all batches with student counts
    batches = Batch.objects.annotate(
        student_count=Count('students'),
        total_revenue=Sum('students__discounted_price'),
        received_payment=Sum('students__advance_payment'),
        pending_payment=Sum('students__second_installment')
    ).values('batch_number', 'student_count', 'total_revenue', 'received_payment', 'pending_payment')
    
    # Convert Decimal objects to float for JSON serialization
    batch_data = []
    for batch in batches:
        batch_data.append({
            'batch_number': batch['batch_number'],
            'student_count': batch['student_count'],
            'total_revenue': float(batch['total_revenue']) if batch['total_revenue'] else 0,
            'received_payment': float(batch['received_payment']) if batch['received_payment'] else 0,
            'pending_payment': float(batch['pending_payment']) if batch['pending_payment'] else 0
        })
    
    return JsonResponse({'batches': batch_data})

# Pending Payment Invoice View
@login_required(login_url='login')
def generate_pending_invoice(request, student_id):
    """Generate printable invoice for a student's pending payment"""
    # Get CSR profile
    try:
        # Get CSR profile directly from the database
        csr_profile = CSRProfile.objects.get(user=request.user)
    except CSRProfile.DoesNotExist:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('login')
    
    # Get student object
    student = get_object_or_404(Student, id=student_id)
    
    # Check if the student has pending payment status
    if student.payment_status != 'pending':
        messages.error(request, f"{student.name} does not have any pending payments.")
        return redirect('student_management')
    
    # Calculate the pending amount
    pending_amount = student.remaining_balance
    
    if pending_amount <= 0:
        messages.warning(request, f"{student.name} has no remaining balance to pay.")
        return redirect('student_management')
    
    # Set due date to current date only if it's currently null or empty
    # This represents when the pending invoice was first generated
    update_fields = ['total_amount']
    
    if student.due_date is None or student.due_date == '':
        student.due_date = timezone.now().date()
        update_fields.append('due_date')
    
    # Update total_amount by adding advance_payment and second_installment
    student.total_amount = student.advance_payment + student.second_installment
    
    # Save the student object with updated fields
    # We don't update second_installment_due_date here as that's set separately
    student.save(update_fields=update_fields)
    
    # Get or create StudentInvoice record for this student
    student_invoice, created = StudentInvoice.objects.get_or_create(student=student)
    
    # Determine which invoice settings will be used (same logic as get_invoice_context)
    invoice_settings = None

    # 1) Try student's creator invoice settings
    if student.created_by:
        invoice_settings = InvoiceSettings.objects.filter(csr=student.created_by).first()

    # 2) If not found, fall back to a Lead CSR with settings
    if not invoice_settings:
        invoice_settings = InvoiceSettings.objects.filter(csr__lead_role=True).first()

    # 3) As a final fallback, use (or create) settings for the logged-in CSR
    if not invoice_settings:
        invoice_settings, _ = InvoiceSettings.objects.get_or_create(
            csr=csr_profile,
            defaults={
                'current_serial_number': 1000
            }
        )

    # Only increment and persist the serial number if this is the first time generating a pending invoice
    if student_invoice.pending_invoice_no == 0:
        # Increment the serial number
        invoice_settings.current_serial_number += 1
        invoice_settings.save()
        
        # Save the invoice number to the student's invoice record
        student_invoice.pending_invoice_no = invoice_settings.current_serial_number
        student_invoice.save()
    
    # Use the utility function to render the invoice with is_pending=True
    return render_printable_invoice(request, student, is_pending=True)

# Report Generation Views
@login_required
def report_students(request):
    """Generate student details report with filters"""
    # Get filter parameters
    batch_id = request.GET.get('batch')
    course_id = request.GET.get('course')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_status = request.GET.get('payment_status')
    export_format = request.GET.get('export')
    
    # Base queryset
    students = Student.objects.all().select_related('batch', 'created_by').prefetch_related('courses')
    
    # Apply filters if provided
    if batch_id:
        students = students.filter(batch_id=batch_id)
    
    if course_id:
        students = students.filter(courses__id=course_id)
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Get students who made payments (advance or second installment) within the date range
        students = students.filter(
            Q(created_at__date__gte=start_date, created_at__date__lte=end_date) |
            Q(due_date__gte=start_date, due_date__lte=end_date)
        ).distinct()
    elif start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        # Get students who made payments on or after this date
        students = students.filter(
            Q(created_at__date__gte=start_date) | Q(due_date__gte=start_date)
        ).distinct()
    elif end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Get students who made payments on or before this date
        students = students.filter(
            Q(created_at__date__lte=end_date) | Q(due_date__lte=end_date)
        ).distinct()
    
    if payment_status:
        students = students.filter(payment_status=payment_status)
    
    # Get all batches and courses for the filter dropdown
    batches = Batch.objects.all()
    courses = Course.objects.all()
    
    # Check if we need to export to Excel
    if export_format == 'excel':
        # Create a new workbook
        wb = Workbook()
        
        # Create Student Details Sheet
        ws = wb.active
        ws.title = "Student Details"
        
        # Define headers for Excel export
        headers = [
            'ID', 'Name', 'Course', 'Status', 'CSR', 'Phone', 'Email', 'Batch',
            'Registration Date', 'Pending Payment Due Date', 'Original Price', 'Discounted Price', 'Advance Payment', 'Second Installment', 'Balance', 'Total Amount',
            'Advance Payment in Range', 'Second Installment in Range', 'Total Payment in Range'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Sort students by date
        students = sorted(students, key=lambda s: s.created_at if s.created_at else timezone.now())
        
        # Write student data
        row = 2
        for index, student in enumerate(students, 1):
            courses_list = ', '.join([course.name for course in student.courses.all()])
            # Use the balance field from the Student model
            balance = student.balance if student.balance is not None else 0
            
            # Calculate payments within date range
            advance_in_range = 0
            second_installment_in_range = 0
            
            if start_date and end_date:
                # Check if advance payment (created_at) falls within range
                if student.created_at and start_date <= student.created_at.date() <= end_date:
                    advance_in_range = float(student.advance_payment) if student.advance_payment else 0
                
                # Check if second installment (due_date) falls within range
                if student.due_date and start_date <= student.due_date <= end_date:
                    second_installment_in_range = float(student.second_installment) if student.second_installment else 0
            elif start_date:
                # Check if payments are on or after start date
                if student.created_at and student.created_at.date() >= start_date:
                    advance_in_range = float(student.advance_payment) if student.advance_payment else 0
                if student.due_date and student.due_date >= start_date:
                    second_installment_in_range = float(student.second_installment) if student.second_installment else 0
            elif end_date:
                # Check if payments are on or before end date
                if student.created_at and student.created_at.date() <= end_date:
                    advance_in_range = float(student.advance_payment) if student.advance_payment else 0
                if student.due_date and student.due_date <= end_date:
                    second_installment_in_range = float(student.second_installment) if student.second_installment else 0
            
            total_payment_in_range = advance_in_range + second_installment_in_range
            
            # Write student details to Excel
            ws.cell(row=row, column=1, value=index)
            ws.cell(row=row, column=2, value=student.name)
            ws.cell(row=row, column=3, value=courses_list)
            ws.cell(row=row, column=4, value='Paid' if student.payment_status == 'paid' else 'Pending')
            ws.cell(row=row, column=5, value=student.get_creator_name() if hasattr(student, 'get_creator_name') else 'N/A')
            ws.cell(row=row, column=6, value=student.phone_number if hasattr(student, 'phone_number') else '')
            ws.cell(row=row, column=7, value=student.email if hasattr(student, 'email') else '')
            ws.cell(row=row, column=8, value=student.batch.batch_number if student.batch else 'N/A')
            ws.cell(row=row, column=9, value=student.created_at.strftime('%Y-%m-%d') if student.created_at else '')
            ws.cell(row=row, column=10, value=student.due_date.strftime('%Y-%m-%d') if student.due_date else '')
            ws.cell(row=row, column=11, value=float(student.total_fees) if student.total_fees else 0)
            ws.cell(row=row, column=12, value=float(student.discounted_price) if student.discounted_price else 0)
            ws.cell(row=row, column=13, value=float(student.advance_payment) if student.advance_payment else 0)
            ws.cell(row=row, column=14, value=float(student.second_installment) if student.second_installment else 0)
            ws.cell(row=row, column=15, value=float(balance))
            ws.cell(row=row, column=16, value=float(student.total_amount) if student.total_amount else 0)
            ws.cell(row=row, column=17, value=advance_in_range)
            ws.cell(row=row, column=18, value=second_installment_in_range)
            ws.cell(row=row, column=19, value=total_payment_in_range)
            
            row += 1
        
        # Auto-adjust column widths for the worksheet
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create the HttpResponse with Excel content
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=student_details_report.xlsx'
        return response
    
    # Get the currently logged-in user's CSR profile if it exists
    csr = None
    if hasattr(request.user, 'csrprofile'):
        csr = request.user.csrprofile
    
    # Determine which template to use based on whether the user is admin or CSR
    if request.user.is_superuser:
        template = 'invoice/report_students.html'
    else:
        template = 'invoice/report_students_csr.html'
    
    # Render the template with context
    context = {
        'students': students,
        'batches': batches,
        'courses': courses,
        'selected_batch': batch_id,
        'selected_course': course_id,
        'start_date': start_date,
        'end_date': end_date,
        'csr': csr,
    }
    
    return render(request, template, context)

@login_required
def report_students_ajax(request):
    """AJAX endpoint for student details report with filters"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    # Get filter parameters
    batch_id = request.GET.get('batch', '')
    course_id = request.GET.get('course', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    payment_status = request.GET.get('payment_status', '')
    
    # Base queryset
    students = Student.objects.all().select_related('batch', 'created_by').prefetch_related('courses')
    
    # Apply filters if provided
    if batch_id:
        students = students.filter(batch_id=batch_id)
    
    if course_id:
        students = students.filter(courses__id=course_id)
    
    # Apply date filtering - only show students who made payments within the date range
    if start_date or end_date:
        date_filter = Q()
        
        if start_date and end_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            # Students who made advance payment within range OR second installment within range
            date_filter = (
                Q(created_at__date__gte=start_date_obj, created_at__date__lte=end_date_obj) |
                Q(due_date__gte=start_date_obj, due_date__lte=end_date_obj)
            )
        elif start_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            # Students who made payments on or after this date
            date_filter = (
                Q(created_at__date__gte=start_date_obj) |
                Q(due_date__gte=start_date_obj)
            )
        elif end_date:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            # Students who made payments on or before this date
            date_filter = (
                Q(created_at__date__lte=end_date_obj) |
                Q(due_date__lte=end_date_obj)
            )
        
        students = students.filter(date_filter).distinct()
    
    if payment_status:
        students = students.filter(payment_status=payment_status)
    
    # Prepare student data for JSON response
    student_list = []
    for student in students:
        courses_list = [course.name for course in student.courses.all()]
        
        # Calculate payments within date range
        advance_in_range = 0
        second_installment_in_range = 0
        
        if start_date and end_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            
            # Check if advance payment (created_at) falls within range
            if student.created_at and start_date_obj <= student.created_at.date() <= end_date_obj:
                advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            
            # Check if second installment (due_date) falls within range
            if student.due_date and start_date_obj <= student.due_date <= end_date_obj:
                second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        elif start_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            
            # Check if payments are on or after start date
            if student.created_at and student.created_at.date() >= start_date_obj:
                advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            if student.due_date and student.due_date >= start_date_obj:
                second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        elif end_date:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            
            # Check if payments are on or before end date
            if student.created_at and student.created_at.date() <= end_date_obj:
                advance_in_range = float(student.advance_payment) if student.advance_payment else 0
            if student.due_date and student.due_date <= end_date_obj:
                second_installment_in_range = float(student.second_installment) if student.second_installment else 0
        
        total_payment_in_range = advance_in_range + second_installment_in_range
        
        student_list.append({
            'id': student.id,
            'name': student.name,
            'courses': courses_list,
            'payment_status': student.payment_status,
            'creator_name': student.get_creator_name() if hasattr(student, 'get_creator_name') else 'N/A',
            'phone_number': student.phone_number,
            'guardian_name': student.guardian_name,
            'batch_number': student.batch.batch_number if student.batch else 'N/A',
            'created_at': student.created_at.strftime('%Y-%m-%d') if student.created_at else '',
            'total_fees': float(student.total_fees) if student.total_fees else 0,
            'discounted_price': float(student.discounted_price) if student.discounted_price else 0,
            'advance_payment': float(student.advance_payment) if student.advance_payment else 0,
            'second_installment': float(student.second_installment) if student.second_installment else 0,
            'balance': float(student.balance) if student.balance is not None else 0,
            'total_amount': float(student.total_amount) if student.total_amount else 0,
            'advance_in_range': advance_in_range,
            'second_installment_in_range': second_installment_in_range,
            'total_payment_in_range': total_payment_in_range,
        })
    
    # Calculate totals for payments within date range
    total_advance_in_range = sum(student['advance_in_range'] for student in student_list)
    total_second_installment_in_range = sum(student['second_installment_in_range'] for student in student_list)
    total_payment_in_range = sum(student['total_payment_in_range'] for student in student_list)
    
    # Prepare response data
    response_data = {
        'students': student_list,
        'student_count': len(student_list),
        'total_advance_in_range': total_advance_in_range,
        'total_second_installment_in_range': total_second_installment_in_range,
        'total_payment_in_range': total_payment_in_range
    }
    
    return JsonResponse(response_data)

@login_required
def report_revenue(request):
    """Generate revenue report by batch and course with filters"""
    # Get filter parameters
    batch_id = request.GET.get('batch')
    course_id = request.GET.get('course')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    export_format = request.GET.get('export')
    
    # Base queryset
    students = Student.objects.all().select_related('batch').prefetch_related('courses')
    
    # Apply filters if provided
    if batch_id:
        students = students.filter(batch_id=batch_id)
    
    if course_id:
        students = students.filter(courses__id=course_id)
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        # Get students who made payments (advance or second installment) within the date range
        students = students.filter(
            Q(created_at__date__gte=start_date, created_at__date__lte=end_date) |
            Q(due_date__gte=start_date, due_date__lte=end_date)
        ).distinct()
    elif start_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        # Get students who made payments on or after this date
        students = students.filter(
            Q(created_at__date__gte=start_date) | Q(due_date__gte=start_date)
        ).distinct()
    elif end_date:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        # Get students who made payments on or before this date
        students = students.filter(
            Q(created_at__date__lte=end_date) | Q(due_date__lte=end_date)
        ).distinct()
    
    # Get all batches and courses for the filter dropdown
    batches = Batch.objects.all()
    courses = Course.objects.all()
    
    # Generate revenue data using date-range-specific payments
    batch_revenue, course_revenue = calculate_date_range_revenue(students, start_date, end_date)
    
    # Calculate totals for batch revenue
    batch_total_revenue = sum(item['total_revenue'] or 0 for item in batch_revenue)
    batch_received_payment = sum(item['received_payment'] or 0 for item in batch_revenue)
    batch_pending_payment = sum(item['pending_payment'] or 0 for item in batch_revenue)
    batch_student_count = sum(item['student_count'] or 0 for item in batch_revenue)
    
    # Calculate totals for course revenue
    course_total_revenue = sum(item['total_revenue'] or 0 for item in course_revenue)
    course_received_payment = sum(item['received_payment'] or 0 for item in course_revenue)
    course_pending_payment = sum(item['pending_payment'] or 0 for item in course_revenue)
    course_student_count = sum(item['student_count'] or 0 for item in course_revenue)
    
    # Check if we need to export
    if export_format == 'excel':
        # Create a new workbook
        wb = Workbook()
        
        # Create Batch Revenue Sheet
        ws_batch = wb.active
        ws_batch.title = "Revenue by Batch"
        
        # Write headers
        headers = ['Batch', 'Total Revenue', 'Received Payment', 'Pending Payment', 'Student Count']
        for col, header in enumerate(headers, 1):
            ws_batch.cell(row=1, column=col, value=header)
        
        # Write batch data
        row = 2
        for item in batch_revenue:
            ws_batch.cell(row=row, column=1, value=item['batch__batch_number'] or 'N/A')
            ws_batch.cell(row=row, column=2, value=float(item['total_revenue'] or 0))
            ws_batch.cell(row=row, column=3, value=float(item['received_payment'] or 0))
            ws_batch.cell(row=row, column=4, value=float(item['pending_payment'] or 0))
            ws_batch.cell(row=row, column=5, value=item['student_count'] or 0)
            row += 1
        
        # Add totals row
        ws_batch.cell(row=row, column=1, value='TOTAL')
        ws_batch.cell(row=row, column=2, value=float(sum(item['total_revenue'] or 0 for item in batch_revenue)))
        ws_batch.cell(row=row, column=3, value=float(sum(item['received_payment'] or 0 for item in batch_revenue)))
        ws_batch.cell(row=row, column=4, value=float(sum(item['pending_payment'] or 0 for item in batch_revenue)))
        ws_batch.cell(row=row, column=5, value=sum(item['student_count'] or 0 for item in batch_revenue))
        
        # Create Course Revenue Sheet
        ws_course = wb.create_sheet(title="Revenue by Course")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws_course.cell(row=1, column=col, value=header.replace('Batch', 'Course'))
        
        # Write course data
        row = 2
        for item in course_revenue:
            ws_course.cell(row=row, column=1, value=item['courses__name'] or 'N/A')
            ws_course.cell(row=row, column=2, value=float(item['total_revenue'] or 0))
            ws_course.cell(row=row, column=3, value=float(item['received_payment'] or 0))
            ws_course.cell(row=row, column=4, value=float(item['pending_payment'] or 0))
            ws_course.cell(row=row, column=5, value=item['student_count'] or 0)
            row += 1
        
        # Add totals row
        ws_course.cell(row=row, column=1, value='TOTAL')
        ws_course.cell(row=row, column=2, value=float(sum(item['total_revenue'] or 0 for item in course_revenue)))
        ws_course.cell(row=row, column=3, value=float(sum(item['received_payment'] or 0 for item in course_revenue)))
        ws_course.cell(row=row, column=4, value=float(sum(item['pending_payment'] or 0 for item in course_revenue)))
        ws_course.cell(row=row, column=5, value=sum(item['student_count'] or 0 for item in course_revenue))
        
        # Auto-adjust column widths for both worksheets
        for worksheet in [ws_batch, ws_course]:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create the HttpResponse with Excel content
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=student_details_report.xlsx'
        return response
    
    # Get the current CSR profile for the sidebar
    try:
        csr = request.user.csr_profile
    except:
        csr = None
    
    # Render the template with filters
    context = {
        'batch_revenue': batch_revenue,
        'course_revenue': course_revenue,
        'batches': batches,
        'courses': courses,
        'selected_batch': batch_id,
        'selected_course': course_id,
        'start_date': start_date,
        'end_date': end_date,
        'csr': csr,
        # Add calculated totals
        'batch_total_revenue': batch_total_revenue,
        'batch_received_payment': batch_received_payment,
        'batch_pending_payment': batch_pending_payment,
        'batch_student_count': batch_student_count,
        'course_total_revenue': course_total_revenue,
        'course_received_payment': course_received_payment,
        'course_pending_payment': course_pending_payment,
        'course_student_count': course_student_count,
    }
    
    if request.user.is_staff:
        return render(request, 'invoice/report_revenue.html', context)
    else:
        return render(request, 'invoice/report_revenue_csr.html', context)


@login_required
def report_revenue_ajax(request):
    """AJAX endpoint for revenue report filtering"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    # Get filter parameters
    batch_id = request.GET.get('batch', '')
    course_id = request.GET.get('course', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Query students with filters
    students = Student.objects.all()
    
    if batch_id:
        students = students.filter(batch_id=batch_id)
    
    if course_id:
        students = students.filter(courses__id=course_id)
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Get students who made payments (advance or second installment) within the date range
        students = students.filter(
            Q(created_at__date__gte=start_date, created_at__date__lte=end_date) |
            Q(due_date__gte=start_date, due_date__lte=end_date)
        ).distinct()
    elif start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        # Get students who made payments on or after this date
        students = students.filter(
            Q(created_at__date__gte=start_date) | Q(due_date__gte=start_date)
        ).distinct()
    elif end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Get students who made payments on or before this date
        students = students.filter(
            Q(created_at__date__lte=end_date) | Q(due_date__lte=end_date)
        ).distinct()
    
    # Generate revenue data using date-range-specific payments
    batch_revenue, course_revenue = calculate_date_range_revenue(students, start_date, end_date)
    
    # Calculate totals for batch revenue
    batch_total_revenue = sum(item['total_revenue'] or 0 for item in batch_revenue)
    batch_received_payment = sum(item['received_payment'] or 0 for item in batch_revenue)
    batch_pending_payment = sum(item['pending_payment'] or 0 for item in batch_revenue)
    batch_student_count = sum(item['student_count'] or 0 for item in batch_revenue)
    
    # Calculate totals for course revenue
    course_total_revenue = sum(item['total_revenue'] or 0 for item in course_revenue)
    course_received_payment = sum(item['received_payment'] or 0 for item in course_revenue)
    course_pending_payment = sum(item['pending_payment'] or 0 for item in course_revenue)
    course_student_count = sum(item['student_count'] or 0 for item in course_revenue)
    
    # Convert Decimal objects to float for JSON serialization
    batch_revenue_list = []
    for item in batch_revenue:
        batch_revenue_list.append({
            'batch__batch_number': item['batch__batch_number'],
            'total_revenue': float(item['total_revenue']) if item['total_revenue'] else 0,
            'received_payment': float(item['received_payment']) if item['received_payment'] else 0,
            'pending_payment': float(item['pending_payment']) if item['pending_payment'] else 0,
            'student_count': item['student_count']
        })
    
    course_revenue_list = []
    for item in course_revenue:
        course_revenue_list.append({
            'courses__name': item['courses__name'],
            'total_revenue': float(item['total_revenue']) if item['total_revenue'] else 0,
            'received_payment': float(item['received_payment']) if item['received_payment'] else 0,
            'pending_payment': float(item['pending_payment']) if item['pending_payment'] else 0,
            'student_count': item['student_count']
        })
    
    # Prepare response data
    response_data = {
        'batch_revenue': batch_revenue_list,
        'course_revenue': course_revenue_list,
        'batch_total_revenue': float(batch_total_revenue),
        'batch_received_payment': float(batch_received_payment),
        'batch_pending_payment': float(batch_pending_payment),
        'batch_student_count': batch_student_count,
        'course_total_revenue': float(course_total_revenue),
        'course_received_payment': float(course_received_payment),
        'course_pending_payment': float(course_pending_payment),
        'course_student_count': course_student_count,
    }
    
    return JsonResponse(response_data)


@login_required
def change_password(request):
    """
    View to handle password changes for both admin and CSR users
    """
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Check if current password is correct
        user = authenticate(username=request.user.username, password=current_password)
        
        if user is None:
            messages.error(request, 'Current password is incorrect.')
            return redirect('change_password')
        
        # Check if new passwords match
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('change_password')
        
        # Check password complexity
        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return redirect('change_password')
        
        # Set the new password
        user.set_password(new_password)
        user.save()
        
        # Update session to prevent user from being logged out
        update_session_auth_hash(request, user)
        
        messages.success(request, 'Your password has been changed successfully.')
        
        # Redirect to appropriate dashboard
        if request.user.is_staff:
            return redirect('admin_dashboard')
        else:
            return redirect('csr_dashboard')
    
    # Get the current CSR profile for the sidebar if user is CSR
    try:
        csr = request.user.csr_profile
    except:
        csr = None
    
    return render(request, 'invoice/change_password.html', {'csr': csr})

@staff_member_required
def admin_batch_management(request):
    """Admin batch management view"""
    # Handle batch creation
    if request.method == 'POST':
        batch_number = request.POST.get('batch_number')
        
        if not batch_number:
            messages.error(request, 'Batch number is required.')
        elif Batch.objects.filter(batch_number=batch_number).exists():
            messages.error(request, 'A batch with this number already exists.')
        else:
            batch = Batch.objects.create(
                batch_number=batch_number,
                created_by=None  # Admin created batches don't need CSR assignment
            )
            messages.success(request, f'Batch {batch_number} created successfully.')
    
    # Get all batches with counts
    batches = Batch.objects.all().order_by('-created_at')
    total_batches = batches.count()
    active_batches = batches.filter(status='active').count()
    
    # Get student counts
    total_students = Student.objects.count()
    current_students = Student.objects.filter(batch__status='active').count()
    
    context = {
        'batches': batches,
        'total_batches': total_batches,
        'active_batches': active_batches,
        'total_students': total_students,
        'current_students': current_students,
    }
    
    return render(request, 'invoice/admin_batch_management.html', context)

@staff_member_required
def update_batch_status(request, batch_id):
    """Update batch status (active/inactive) - Admin only"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    try:
        # Get the batch
        batch = get_object_or_404(Batch, id=batch_id)
        
        # Parse the request body
        data = json.loads(request.body)
        new_status = data.get('status')
        
        # Validate status
        if new_status not in ['active', 'inactive']:
            return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
        
        # Update batch status
        batch.status = new_status
        batch.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Batch {batch.batch_number} marked as {new_status}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@staff_member_required
def batch_management(request):
    """Batch management view - Admin only"""
    if request.method == 'POST':
        # Handle adding new batch
        batch_number = request.POST.get('batch_number')
        if batch_number:
            try:
                batch = Batch.objects.create(
                    batch_number=batch_number,
                    created_by=None  # Admin created batches don't have CSR assignment
                )
                messages.success(request, f'Batch {batch_number} created successfully.')
            except Exception as e:
                messages.error(request, f'Error creating batch: {str(e)}')
        return redirect('batch_management')
    
    # Get all batches for display
    batches = Batch.objects.all().order_by('-created_at')
    
    context = {
        'batches': batches,
    }
    
    return render(request, 'invoice/batch_management.html', context)
def commission_report(request):
    """Commission report view for admin dashboard"""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('login')
    
    # Get filter parameters
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    commission_percent = request.GET.get('commission_percent', 1)  # Default 1%
    
    try:
        commission_percent = float(commission_percent)
    except ValueError:
        commission_percent = 1.0
    
    # Get all CSRs for the cards display with student counts
    csrs = CSRProfile.objects.all().order_by('user__first_name', 'user__last_name')
    
    # Add completed student count for each CSR
    for csr in csrs:
        csr.completed_students_count = Student.objects.filter(
            created_by=csr,
            balance=0,
            payment_status='paid'
        ).count()
    
    # Initialize commission data
    commission_data = []
    total_commission = 0
    total_admissions = 0
    total_revenue = 0
    total_payment_in_range = 0
    
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Get students with completed payments (balance = 0) within date range
            completed_students = Student.objects.filter(
                balance=0,  # Only completed payments
                payment_status='paid',  # Payment status is paid
                created_at__date__range=[start_dt, end_dt]
            ).select_related('created_by', 'batch').prefetch_related('courses')
            
            # Calculate commission for each CSR
            for csr in csrs:
                csr_students = completed_students.filter(created_by=csr)
                
                if csr_students.exists():
                    # Calculate total revenue and commission
                    csr_total_revenue = sum(student.discounted_price for student in csr_students)
                    csr_commission = (csr_total_revenue * commission_percent) / 100
                    csr_admissions = csr_students.count()
                    
                    # Calculate total payment received in range for this CSR's students
                    csr_total_payment_in_range = 0
                    
                    # Add individual commission amounts to each student
                    for student in csr_students:
                        student.commission_amount = (student.discounted_price * commission_percent) / 100
                        # Calculate payment received in range (advance + second installment if within range)
                        payment_in_range = 0
                        if start_dt <= student.created_at.date() <= end_dt:
                            payment_in_range += student.advance_payment or 0
                        if student.due_date and start_dt <= student.due_date <= end_dt:
                            payment_in_range += student.second_installment or 0
                        csr_total_payment_in_range += payment_in_range
                    
                    commission_data.append({
                        'csr': csr,
                        'total_revenue': csr_total_revenue,
                        'commission': csr_commission,
                        'admissions': csr_admissions,
                        'students': csr_students,
                        'total_payment_in_range': csr_total_payment_in_range
                    })
                    
                    total_commission += csr_commission
                    total_admissions += csr_admissions
                    total_revenue += csr_total_revenue
                    total_payment_in_range += csr_total_payment_in_range
            
            # Sort by commission amount (highest first)
            commission_data.sort(key=lambda x: x['commission'], reverse=True)
            
        except ValueError as e:
            # Invalid date format
            messages.error(request, f"Invalid date format: {e}")
        except Exception as e:
            # General error handling
            messages.error(request, f"Error calculating commissions: {e}")
    
    context = {
        'csrs': csrs,
        'commission_data': commission_data,
        'total_commission': total_commission,
        'total_admissions': total_admissions,
        'total_revenue': total_revenue,
        'total_payment_in_range': total_payment_in_range,
        'commission_percent': commission_percent,
        'start_date': start_date,
        'end_date': end_date,
        'has_filters': bool(start_date and end_date)
    }
    
    # Debug information (remove in production)
    print(f"Debug: Found {len(csrs)} CSRs")
    for csr in csrs:
        print(f"Debug: CSR {csr.get_full_name()} has {csr.completed_students_count} completed students")
    
    return render(request, 'invoice/commission.html', context)
