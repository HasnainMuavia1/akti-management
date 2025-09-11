from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.db import models
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import csv
from datetime import datetime, timedelta
import pytz

from .models import Trainer, TrainerCourse, Lecture, Attendance, AttendanceReport
from .forms import (
    TrainerCreationForm, TrainerCourseAssignmentForm, LectureForm, 
    AttendanceForm, BulkAttendanceForm, CourseFilterForm, BatchFilterForm,
    TrainerEditForm
)
from example.models import Course, Student, Batch
def _renumber_lectures_for_assignment(trainer_course: TrainerCourse) -> None:
    """Ensure lecture_number reflects chronological order for a given trainer_course.
    Performs a two-phase renumber to avoid unique_together collisions.
    """
    # Phase 1: assign temporary numbers to avoid unique clashes
    lectures = list(Lecture.objects.filter(trainer_course=trainer_course).order_by('date', 'start_time', 'id'))
    # Temporary offset so we don't clash with existing 1..N
    temp_offset = 1000
    for idx, lec in enumerate(lectures, start=1):
        temp_num = temp_offset + idx
        if lec.lecture_number != temp_num:
            Lecture.objects.filter(id=lec.id).update(lecture_number=temp_num)
    # Phase 2: set final sequential numbers
    for idx, lec in enumerate(lectures, start=1):
        Lecture.objects.filter(id=lec.id).update(lecture_number=idx)



def _students_for_assignment(course, batch, schedule):
    qs = Student.objects.filter(courses=course)
    if batch:
        qs = qs.filter(batch=batch)
    if schedule:
        qs = qs.filter(schedule=schedule)
    return qs


def is_admin(user):
    """Check if user is admin"""
    return user.is_authenticated and user.is_staff


def is_trainer(user):
    """Check if user is a trainer"""
    return user.is_authenticated and hasattr(user, 'trainer_profile')


def portal_login(request):
    """Portal login view"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_staff or hasattr(user, 'trainer_profile'):
                login(request, user)
                return redirect('portal:portal_dashboard')
            else:
                messages.error(request, "Access denied. You don't have permission to access the portal.")
        else:
            messages.error(request, "Invalid username or password.")
    
    return render(request, 'portal/login.html')


def portal_logout(request):
    """Portal logout view"""
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('portal:portal_login')


@login_required
def portal_dashboard(request):
    """Main portal dashboard - redirects based on user type"""
    if is_admin(request.user):
        return redirect('portal:admin_dashboard')
    elif is_trainer(request.user):
        return redirect('portal:trainer_dashboard')
    else:
        messages.error(request, "Access denied. You don't have permission to access the portal.")
        return redirect('portal:portal_login')


@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Admin dashboard for attendance management"""
    context = {
        'total_trainers': Trainer.objects.filter(is_active=True).count(),
        'total_courses': Course.objects.count(),
        'total_students': Student.objects.count(),
        'total_batches': Batch.objects.count(),
        'recent_assignments': TrainerCourse.objects.select_related('trainer', 'course').order_by('-assigned_at')[:5],
        'recent_lectures': Lecture.objects.select_related('trainer_course__trainer', 'trainer_course__course').order_by('-created_at')[:5],
    }
    return render(request, 'portal/admin/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def trainer_management(request):
    """Admin view for managing trainers"""
    trainers = Trainer.objects.select_related('user').all()
    
    # Remove inline creation; use dedicated add page instead
    form = TrainerCreationForm()
    
    context = {
        'trainers': trainers,
        'form': form,
    }
    return render(request, 'portal/admin/trainer_management.html', context)


@login_required
@user_passes_test(is_admin)
def trainer_add(request):
    """Dedicated page to add a new trainer (replaces modal)."""
    if request.method == 'POST':
        form = TrainerCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Trainer created successfully!')
            return redirect('portal:trainer_management')
    else:
        form = TrainerCreationForm()
    return render(request, 'portal/admin/trainer_add.html', { 'form': form })


@login_required
@user_passes_test(is_admin)
def trainer_edit(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    user = trainer.user

    if request.method == 'POST':
        form = TrainerEditForm(request.POST, instance=user, trainer=trainer)
        if form.is_valid():
            form.save(trainer=trainer)
            messages.success(request, 'Trainer updated successfully!')
            return redirect('portal:trainer_management')
    else:
        form = TrainerEditForm(instance=user, trainer=trainer)

    context = {
        'trainer': trainer,
        'form': form,
    }
    return render(request, 'portal/admin/trainer_edit.html', context)


@login_required
@user_passes_test(is_admin)
def trainer_delete(request, trainer_id):
    """Delete a trainer and the linked user. Supports AJAX (JSON)."""
    trainer = get_object_or_404(Trainer, id=trainer_id)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)
        messages.error(request, 'Invalid request method')
        return redirect('portal:trainer_management')

    try:
        user = trainer.user
        trainer.delete()  # cascades OneToOne only removes trainer, keep user? we want to remove both
        user.delete()
        if is_ajax:
            return JsonResponse({'success': True})
        messages.success(request, 'Trainer deleted successfully!')
        return redirect('portal:trainer_management')
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        messages.error(request, f'Error deleting trainer: {e}')
        return redirect('portal:trainer_management')


@login_required
@user_passes_test(is_admin)
def course_assignment(request):
    """Admin view for listing course-trainer assignments"""
    assignments = TrainerCourse.objects.select_related('trainer', 'course', 'batch').all()
    total_assignments = assignments.count()
    active_assignments = assignments.filter(is_active=True).count()
    trainers_with_courses = assignments.values('trainer').distinct().count()
    courses_assigned = assignments.values('course').distinct().count()
    context = {
        'assignments': assignments,
        'total_assignments': total_assignments,
        'active_assignments': active_assignments,
        'trainers_with_courses': trainers_with_courses,
        'courses_assigned': courses_assigned,
    }
    return render(request, 'portal/admin/course_assignment.html', context)


@login_required
@user_passes_test(is_admin)
def course_assignment_form(request):
    """Admin view to assign courses to trainers (separate page)."""
    edit_id = request.GET.get('edit')
    edit_instance = None
    if edit_id:
        edit_instance = get_object_or_404(TrainerCourse, id=edit_id)
    if request.method == 'POST':
        if request.POST.get('edit_id'):
            # Update existing assignment (single course)
            assignment = get_object_or_404(TrainerCourse, id=request.POST.get('edit_id'))
            form = TrainerCourseAssignmentForm(request.POST)
            if form.is_valid():
                assignment.trainer = form.cleaned_data['trainer']
                assignment.batch = form.cleaned_data['batch']
                # If multiple courses selected, update the first and create others
                courses = form.cleaned_data['courses']
                if courses:
                    assignment.course = courses[0]
                assignment.save()
                # Create remaining courses for same trainer/batch
                for course in courses[1:]:
                    TrainerCourse.objects.get_or_create(trainer=assignment.trainer, batch=assignment.batch, course=course)
                messages.success(request, 'Assignment updated successfully!')
                return redirect('portal:course_assignment')
        else:
            form = TrainerCourseAssignmentForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Courses assigned successfully!')
                return redirect('portal:course_assignment')
    else:
        if edit_instance:
            # Prepopulate form with current values
            initial = {
                'trainer': edit_instance.trainer,
                'batch': edit_instance.batch,
                'courses': [edit_instance.course.id],
            }
            form = TrainerCourseAssignmentForm(initial=initial)
        else:
            form = TrainerCourseAssignmentForm()
    context = {
        'form': form,
        'edit_instance': edit_instance,
    }
    return render(request, 'portal/admin/course_assignment_form.html', context)


@login_required
@user_passes_test(is_admin)
def delete_assignment(request, assignment_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)
    assignment = get_object_or_404(TrainerCourse, id=assignment_id)
    try:
        assignment.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@user_passes_test(is_admin)
def individual_attendance(request):
    """Admin view for viewing individual student attendance"""
    students = Student.objects.select_related('batch').all()
    selected_student = None
    attendances = []
    
    if request.GET.get('student_id'):
        selected_student = get_object_or_404(Student, id=request.GET.get('student_id'))
        attendances = Attendance.objects.filter(
            student=selected_student
        ).select_related('lecture__trainer_course__course', 'marked_by').order_by('-lecture__date')
    
    # For filters
    all_batches = Batch.objects.all()
    all_courses = Course.objects.all()
    
    context = {
        'students': students,
        'selected_student': selected_student,
        'attendances': attendances,
        'all_batches': all_batches,
        'all_courses': all_courses,
    }
    return render(request, 'portal/admin/individual_attendance.html', context)


@login_required
@user_passes_test(is_admin)
def course_attendance_report(request):
    """Admin view for course attendance reports"""
    form = CourseFilterForm(request.GET or None)
    attendances = []
    
    if form.is_valid() and form.cleaned_data.get('course'):
        course = form.cleaned_data['course']
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        
        # Get all lectures for this course
        lectures = Lecture.objects.filter(trainer_course__course=course)
        if date_from:
            lectures = lectures.filter(date__gte=date_from)
        if date_to:
            lectures = lectures.filter(date__lte=date_to)
        
        attendances = Attendance.objects.filter(
            lecture__in=lectures
        ).select_related('student', 'lecture', 'marked_by')
    
    # Additional context for template: active batches and their courses
    batches = Batch.objects.all()
    courses_by_batch = {}
    for b in batches:
        courses_by_batch[b] = Course.objects.filter(students__batch=b).distinct()
    
    context = {
        'form': form,
        'attendances': attendances,
        'batches': batches,
        'courses_by_batch': courses_by_batch,
    }
    return render(request, 'portal/admin/course_attendance_report.html', context)


@login_required
@user_passes_test(is_admin)
def batch_attendance_report(request):
    """Admin view for batch attendance reports"""
    form = BatchFilterForm(request.GET or None)
    attendances = []
    
    if form.is_valid() and form.cleaned_data.get('batch'):
        batch = form.cleaned_data['batch']
        attendances = Attendance.objects.filter(
            student__batch=batch
        ).select_related('student', 'lecture__trainer_course__course', 'marked_by')
    
    # Additional context for trainers and their assigned courses
    trainers = Trainer.objects.select_related('user').all()
    courses_by_trainer = {}
    for t in trainers:
        courses_by_trainer[t] = TrainerCourse.objects.filter(trainer=t, is_active=True).select_related('course', 'batch')
    
    context = {
        'form': form,
        'attendances': attendances,
        'trainers': trainers,
        'courses_by_trainer': courses_by_trainer,
    }
    return render(request, 'portal/admin/batch_attendance_report.html', context)


@login_required
@user_passes_test(is_trainer)
def trainer_dashboard(request):
    """Trainer dashboard"""
    trainer = request.user.trainer_profile
    assigned_courses = TrainerCourse.objects.filter(trainer=trainer, is_active=True).select_related('course', 'batch')
    # Total students respects batch and schedule if set on assignment
    total_students = 0
    for tc in assigned_courses:
        schedule = getattr(tc, 'schedule', None)
        students_qs = _students_for_assignment(tc.course, tc.batch, schedule)
        total_students += students_qs.count()
    # Today's completed lectures for this trainer
    today = timezone.now().date()
    todays_lectures = Lecture.objects.filter(trainer_course__trainer=trainer, date=today, attendances__isnull=False).distinct().count()
    context = {
        'trainer': trainer,
        'assigned_courses': assigned_courses,
        'total_courses': assigned_courses.count(),
        'total_students': total_students,
        'todays_lectures': todays_lectures,
    }
    return render(request, 'portal/trainer/dashboard.html', context)


@login_required
@user_passes_test(is_trainer)
def trainer_course_detail(request, trainer_course_id):
    """Trainer view for specific course details"""
    trainer = request.user.trainer_profile
    trainer_course = get_object_or_404(TrainerCourse, id=trainer_course_id, trainer=trainer)
    
    # Show only lectures with attendance, sorted by date then lecture number
    lectures = Lecture.objects.filter(trainer_course=trainer_course, attendances__isnull=False).distinct().order_by('date', 'lecture_number')
    
    # Get students enrolled in this course (respect batch and schedule)
    schedule = getattr(trainer_course, 'schedule', None)
    students = _students_for_assignment(trainer_course.course, trainer_course.batch, schedule)
    
    context = {
        'trainer_course': trainer_course,
        'lectures': lectures,
        'students': students,
    }
    return render(request, 'portal/trainer/course_detail.html', context)


@login_required
@user_passes_test(is_trainer)
def trainer_create_attendance_for_date(request, trainer_course_id):
    """Create a lecture for a specific date and redirect to mark attendance."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)

    trainer = request.user.trainer_profile
    trainer_course = get_object_or_404(TrainerCourse, id=trainer_course_id, trainer=trainer)

    try:
        payload = json.loads(request.body or '{}')
    except Exception:
        payload = request.POST

    date_str = payload.get('date')
    if not date_str:
        return JsonResponse({'success': False, 'message': 'Missing date'}, status=400)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Determine next lecture number for that course regardless of date
    last_lecture = Lecture.objects.filter(trainer_course=trainer_course).order_by('-lecture_number').first()
    next_number = 1 if not last_lecture else last_lecture.lecture_number + 1

    # Use Pakistan time for times
    pk_tz = pytz.timezone('Asia/Karachi')
    now_local = timezone.now().astimezone(pk_tz)
    # Start at 10:00 AM local if creating for past date; otherwise now
    default_start = now_local.replace(hour=10, minute=0, second=0, microsecond=0)
    if target_date == now_local.date():
        start_dt = now_local
    else:
        start_dt = pk_tz.localize(datetime.combine(target_date, default_start.timetz()))
    duration_minutes = 60 if ('1_month' in (trainer_course.course.duration or '')) else 90
    end_dt = start_dt + timedelta(minutes=duration_minutes)

    lecture = Lecture.objects.create(
        trainer_course=trainer_course,
        lecture_number=next_number,
        date=target_date,
        start_time=start_dt.time().replace(second=0, microsecond=0),
        end_time=end_dt.time().replace(second=0, microsecond=0),
    )

    return JsonResponse({'success': True, 'redirect': f"{request.build_absolute_uri('').rstrip('/')}{reverse('portal:mark_attendance', args=[lecture.id])}"})


@login_required
@user_passes_test(is_trainer)
def trainer_start_attendance(request, trainer_course_id):
    """Create or resume the next lecture for the trainer's course and redirect to mark_attendance."""
    trainer = request.user.trainer_profile
    trainer_course = get_object_or_404(TrainerCourse, id=trainer_course_id, trainer=trainer)

    # Stop if course is already completed
    if trainer_course.completed_lectures >= trainer_course.total_lectures:
        messages.info(request, 'This course is completed. No further attendance can be marked.')
        return redirect('portal:trainer_course_detail', trainer_course_id=trainer_course.id)

    # Determine next lecture number
    last_lecture = Lecture.objects.filter(trainer_course=trainer_course).order_by('-lecture_number').first()
    next_number = 1 if not last_lecture else last_lecture.lecture_number + 1

    # Cooldown removed: allow opening mark page; time rules enforced on save

    # If the last lecture exists and is not completed, reuse it; otherwise create next
    if last_lecture and not last_lecture.attendances.exists():
        lecture = last_lecture
    else:
        today = timezone.now().date()
        start_dt = timezone.now()
        # Duration: default 90 minutes; for 1 month courses use 60
        duration_minutes = 60 if ('1_month' in (trainer_course.course.duration or '')) else 90
        end_dt = start_dt + timedelta(minutes=duration_minutes)
        lecture = Lecture.objects.create(
            trainer_course=trainer_course,
            lecture_number=next_number,
            date=today,
            start_time=start_dt.time().replace(second=0, microsecond=0),
            end_time=end_dt.time().replace(second=0, microsecond=0),
        )
    # Ensure numbering is chronological after any creation
    _renumber_lectures_for_assignment(trainer_course)
    return redirect('portal:mark_attendance', lecture_id=lecture.id)


@login_required
@user_passes_test(is_trainer)
def mark_attendance(request, lecture_id):
    """Trainer view for marking attendance"""
    trainer = request.user.trainer_profile
    lecture = get_object_or_404(Lecture, id=lecture_id, trainer_course__trainer=trainer)
    
    if request.method == 'POST':
        # Handle bulk attendance submission
        data = json.loads(request.body)
        # If a specific date is provided, set the lecture's date accordingly
        selected_date_str = data.get('date')
        selected_date = None
        if selected_date_str:
            try:
                selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            except ValueError:
                selected_date = None

        # Enforce slot rule: if marking for today and current time is before slot end, block save
        pk_tz = pytz.timezone('Asia/Karachi')
        now_local = timezone.now().astimezone(pk_tz)
        target_date = selected_date or lecture.date
        if target_date == now_local.date():
            # Determine slot end
            duration_minutes = 60 if ('1_month' in (lecture.trainer_course.course.duration or '')) else 90
            if lecture.end_time:
                end_dt = pk_tz.localize(datetime.combine(target_date, lecture.end_time))
            else:
                # Fallback: compute from start_time or assume now + duration
                base_start = lecture.start_time if lecture.start_time else now_local.time()
                base_start_dt = pk_tz.localize(datetime.combine(target_date, base_start))
                end_dt = base_start_dt + timedelta(minutes=duration_minutes)
            # Also block if any other ongoing lecture (including this one) hasn't ended yet today
            ongoing_block = False
            candidate_lectures = Lecture.objects.filter(
                trainer_course=lecture.trainer_course,
                date=target_date,
            )
            for lec in candidate_lectures:
                if lec.end_time:
                    lec_end = pk_tz.localize(datetime.combine(target_date, lec.end_time))
                else:
                    lec_start_time = lec.start_time if lec.start_time else now_local.time()
                    lec_start = pk_tz.localize(datetime.combine(target_date, lec_start_time))
                    lec_end = lec_start + timedelta(minutes=duration_minutes)
                # Consider it ongoing if start <= now < end
                lec_start_cmp = lec.start_time or base_start
                lec_start_dt_cmp = pk_tz.localize(datetime.combine(target_date, lec_start_cmp))
                if (now_local >= lec_start_dt_cmp) and (now_local < lec_end):
                    ongoing_block = True
                    end_dt = max(end_dt, lec_end)
                    break

            if now_local < end_dt or ongoing_block:
                # Cleanup: avoid leaving behind empty lectures started in this slot
                if Attendance.objects.filter(lecture=lecture).count() == 0:
                    try:
                        # Only delete if no attendance saved yet
                        lecture.delete()
                        # Note: caller page may hold a stale lecture_id; front-end stays put and shows error toast
                    except Exception:
                        pass
                return JsonResponse({
                    'success': False,
                    'message': f"You can submit attendance after {end_dt.strftime('%I:%M %p')}"
                }, status=400)

        # If this is the first time marking for this lecture, set start/end based on now but anchor date to selected (if any)
        if not Attendance.objects.filter(lecture=lecture).exists():
            pk_tz = pytz.timezone('Asia/Karachi')
            now = timezone.now().astimezone(pk_tz)
            duration_minutes = 60 if ('1_month' in (lecture.trainer_course.course.duration or '')) else 90
            new_end = now + timedelta(minutes=duration_minutes)
            lecture.date = selected_date or now.date()
            lecture.start_time = now.time().replace(second=0, microsecond=0)
            lecture.end_time = new_end.time().replace(second=0, microsecond=0)
            lecture.save(update_fields=['date', 'start_time', 'end_time'])
            _renumber_lectures_for_assignment(lecture.trainer_course)
        elif selected_date:
            # If attendance exists but user changed date, update date only
            lecture.date = selected_date
            lecture.save(update_fields=['date'])
            _renumber_lectures_for_assignment(lecture.trainer_course)
        success_count = 0
        
        for student_data in data.get('attendances', []):
            student_id = student_data.get('student_id')
            status = student_data.get('status')
            
            if student_id and status:
                student = get_object_or_404(Student, id=student_id)
                attendance, created = Attendance.objects.get_or_create(
                    lecture=lecture,
                    student=student,
                    defaults={
                        'status': status,
                        'marked_by': trainer
                    }
                )
                
                if not created:
                    attendance.status = status
                    attendance.marked_by = trainer
                    attendance.save()
                
                success_count += 1
        
        # No is_completed flag anymore
        
        return JsonResponse({
            'success': True,
            'message': f'Attendance marked for {success_count} students',
            'success_count': success_count
        })
    
    # Get students enrolled in this course (respect batch)
    schedule = getattr(lecture.trainer_course, 'schedule', None)
    students = _students_for_assignment(lecture.trainer_course.course, lecture.trainer_course.batch, schedule).select_related('batch')
    
    # Get existing attendance records
    existing_attendance = {}
    attendance_counts = {"present": 0, "absent": 0}
    for attendance in Attendance.objects.filter(lecture=lecture):
        existing_attendance[attendance.student.id] = {
            'status': attendance.status
        }
        if attendance.status in attendance_counts:
            attendance_counts[attendance.status] += 1
    
    # Build per-student attendance history for this course
    attendance_history = {}
    for student in students:
        records = Attendance.objects.filter(
            student=student,
            lecture__trainer_course__course=lecture.trainer_course.course
        ).select_related('lecture').order_by('-lecture__date')
        attendance_history[student.id] = list(records)

    # Prepare previous lectures (exclude current) for horizontal history columns (respect batch)
    # Only get lectures from the same trainer course assignment (same batch and schedule)
    prev_lectures = Lecture.objects.filter(
        trainer_course=lecture.trainer_course
    ).exclude(id=lecture.id).order_by('date', 'lecture_number')
    prev_dates = [lec.date for lec in prev_lectures]
    
    # Additional safety check: ensure all previous lectures are from the same assignment
    if prev_lectures:
        # Verify all lectures belong to the same trainer course assignment
        for lec in prev_lectures:
            if lec.trainer_course_id != lecture.trainer_course_id:
                print(f"WARNING: Lecture {lec.id} belongs to different trainer course assignment")
                print(f"Expected: {lecture.trainer_course_id}, Got: {lec.trainer_course_id}")

    # Map: student_id -> {date: 'P'|'A'} - ONLY for students in current assignment
    prev_status_by_student = {s.id: {} for s in students}
    if prev_lectures:
        # Only get attendance records for students in the current assignment
        prev_atts = Attendance.objects.filter(
            lecture__in=prev_lectures,
            student__in=students  # Filter by current assignment students only
        ).select_related('lecture', 'student')
        
        # Debug: Log what we found
        # Debug logs retained intentionally for troubleshooting
        
        for att in prev_atts:
            if att.student_id in prev_status_by_student:  # Double-check student is in current assignment
                # Use date + lecture_number as key to distinguish multiple lectures on same day
                key = f"{att.lecture.date}_{att.lecture.lecture_number}"
                prev_status_by_student[att.student_id][key] = 'P' if att.status != 'absent' else 'A'
                print(f"Student {att.student.name} on {att.lecture.date} Lecture {att.lecture.lecture_number}: {att.status}")
        
        print(f"Final prev_status_by_student: {prev_status_by_student}")
    
    context = {
        'lecture': lecture,
        'students': students,
        'existing_attendance': existing_attendance,
        'attendance_counts': attendance_counts,
        'attendance_history': attendance_history,
        'prev_lectures': prev_lectures,
        'prev_dates': prev_dates,
        'prev_status_by_student': prev_status_by_student,
    }
    return render(request, 'portal/trainer/mark_attendance.html', context)


@login_required
@user_passes_test(is_trainer)
def trainer_reports(request):
    """Trainer view for generating reports"""
    trainer = request.user.trainer_profile
    assigned_courses = TrainerCourse.objects.filter(trainer=trainer, is_active=True).select_related('course', 'batch')
    # Build students count per assignment (respect batch and schedule if set)
    students_per_assignment = {}
    for tc in assigned_courses:
        schedule = getattr(tc, 'schedule', None)
        students_per_assignment[tc.id] = _students_for_assignment(tc.course, tc.batch, schedule).count()
    context = {
        'assigned_courses': assigned_courses,
        'students_per_assignment': students_per_assignment,
    }
    return render(request, 'portal/trainer/reports.html', context)


# AJAX Views
@csrf_exempt
@require_http_methods(["POST"])
def ajax_mark_attendance(request):
    """AJAX endpoint for marking attendance"""
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Authentication required'})
    
    try:
        data = json.loads(request.body)
        lecture_id = data.get('lecture_id')
        student_id = data.get('student_id')
        status = data.get('status')
        remarks = data.get('remarks', '')
        
        if not all([lecture_id, student_id, status]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'})
        
        lecture = get_object_or_404(Lecture, id=lecture_id)
        student = get_object_or_404(Student, id=student_id)
        
        # Check if trainer has permission
        if not hasattr(request.user, 'trainer_profile') or lecture.trainer_course.trainer != request.user.trainer_profile:
            return JsonResponse({'success': False, 'message': 'Permission denied'})
        
        attendance, created = Attendance.objects.get_or_create(
            lecture=lecture,
            student=student,
            defaults={
                'status': status,
                'remarks': remarks,
                'marked_by': request.user.trainer_profile
            }
        )
        
        if not created:
            attendance.status = status
            attendance.remarks = remarks
            attendance.marked_by = request.user.trainer_profile
            attendance.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Attendance marked successfully',
            'created': created
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def download_attendance_report(request, report_type, object_id=None):
    """Download attendance report with unified layout.
    Layout (XLSX):
    - Columns: Student Name | Course | Lecture #N (merged) with subcolumns: Day | Date | ... repeated per lecture | Marked By
    - Values: 'present' / 'absent' (lowercase) under merged Day/Date value cells
    - trainer: multi-sheet (one sheet per course; sheet: Course - Batch)
    - course: single sheet (file: Course - Batch)
    - batch: multi-sheet (one per course with attendance in that batch; sheet: Course - Trainer)
    - admin (all): multi-sheet (one per active course)
    Supports optional ?batch=<id> to filter lectures/students for course export.
    """
    if not (is_admin(request.user) or is_trainer(request.user)):
        messages.error(request, "Permission denied")
        return redirect('portal:portal_dashboard')

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font
    except Exception:
        return JsonResponse({'success': False, 'message': 'openpyxl is required to export Excel. Please install: pip install openpyxl'})

    def sanitize_title(raw: str) -> str:
        title = raw
        for ch in ['/', '\\', '?', '*', '[', ']', ':']:
            title = title.replace(ch, '-')
        return title[:31]

    def _students_for_assignment(course, batch, schedule):
        """Helper to get students for a given course, batch, and optional schedule."""
        if batch:
            students_qs = Student.objects.filter(courses=course, batch=batch)
        else:
            students_qs = Student.objects.filter(courses=course)

        if schedule:
            # Filter students by their schedule field (weekdays/weekend)
            students_qs = students_qs.filter(schedule=schedule)
        
        return students_qs

    def build_sheet_for_course(wb, course_obj, lectures_qs, students_qs, title):
        # Order lectures chronologically and keep only those with any attendance
        lectures = list(lectures_qs.order_by('date', 'lecture_number').filter(attendances__isnull=False).distinct())
        if not lectures:
            return False
        ws = wb.create_sheet(title=sanitize_title(title))
        # Two header rows
        header_row1 = ['Student Name', 'Course']
        header_row2 = ['', '']
        for idx, lec in enumerate(lectures, start=1):
            header_row1.extend([f"Lecture # {idx}", ''])
            date_str = lec.date.strftime('%d/%m/%Y') if lec.date else ''
            header_row2.extend([lec.date.strftime('%a').upper() if lec.date else '', date_str])
        header_row1.append('Marked By')
        header_row2.append('')
        ws.append(header_row1)
        ws.append(header_row2)
        # Merge lecture headers
        center_bold = Alignment(horizontal='center', vertical='center')
        center = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)
        ws.freeze_panes = 'C3'
        col = 3  # starting column C (after Student Name, Course)
        for _ in lectures:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            col += 2
        for row_idx in (1, 2):
            for cell in ws[row_idx]:
                cell.alignment = center_bold
                cell.font = bold_font
        # Attendance map
        attendances = Attendance.objects.filter(lecture__in=lectures, student__in=students_qs).select_related('lecture', 'student')
        att_map = {}
        for a in attendances:
            symbol = 'present' if a.status == 'present' else ('absent' if a.status == 'absent' else '-')
            att_map[(a.student_id, a.lecture_id)] = symbol
        # Marked By (trainer) resolution
        trainer_names = set()
        for lec in lectures:
            if lec.trainer_course and lec.trainer_course.trainer:
                trainer_names.add(lec.trainer_course.trainer.name)
        marked_by_value = next(iter(trainer_names)) if len(trainer_names) == 1 else ('Multiple' if len(trainer_names) > 1 else '')
        for s in students_qs.order_by('name'):
            current_row = ws.max_row + 1
            ws.cell(row=current_row, column=1, value=s.name)
            ws.cell(row=current_row, column=2, value=course_obj.name)
            start_col = 3
            for lec in lectures:
                sym = att_map.get((s.id, lec.id), '-')
                ws.cell(row=current_row, column=start_col, value=sym).alignment = center
                ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col+1)
                start_col += 2
            ws.cell(row=current_row, column=start_col, value=marked_by_value)
        # Autosize
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = '' if cell.value is None else str(cell.value)
                    if len(val) > max_len:
                        max_len = len(val)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 26)
        return True

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    filename = 'attendance.xlsx'
    sheets_created = 0

    # Course-specific export (optional batch filter via query param)
    if report_type == 'course' and object_id:
        course = get_object_or_404(Course, id=object_id)
        lectures = Lecture.objects.filter(trainer_course__course=course).select_related('trainer_course__batch', 'trainer_course__trainer')
        batch_id = request.GET.get('batch') or request.GET.get('batch_id')
        schedule = request.GET.get('schedule')  # optional explicit schedule
        if batch_id:
            lectures = lectures.filter(trainer_course__batch_id=batch_id)
            batch_obj = Batch.objects.get(id=batch_id)
            students = _students_for_assignment(course, batch_obj, schedule)
            batch_number = batch_obj.batch_number
            sheet_title = f"{course.name} - Batch {batch_number}"
            filename = f"{sanitize_title(course.name)} - Batch {sanitize_title(batch_number)}.xlsx"
        else:
            students = _students_for_assignment(course, None, schedule)
            sheet_title = course.name
            filename = f"{sanitize_title(course.name)}.xlsx"
        if build_sheet_for_course(wb, course, lectures, students, sheet_title):
            sheets_created += 1
    
    # Trainer course-specific export: trainer downloads one specific course
    elif report_type == 'trainer_course' and object_id and is_trainer(request.user):
        trainer = request.user.trainer_profile
        trainer_course = get_object_or_404(TrainerCourse, id=object_id, trainer=trainer, is_active=True)
        
        # Only include lectures for this specific trainer assignment
        lectures = Lecture.objects.filter(trainer_course=trainer_course).select_related('trainer_course__batch')
        schedule = getattr(trainer_course, 'schedule', None)
        # Filter students by the trainer's specific assignment (batch and schedule)
        students = _students_for_assignment(trainer_course.course, trainer_course.batch, schedule)
        
        if trainer_course.batch_id:
            sheet_title = f"{trainer_course.course.name} - {trainer_course.batch.batch_number}"
            filename = f"{sanitize_title(trainer_course.course.name)} - {sanitize_title(trainer_course.batch.batch_number)}.xlsx"
        else:
            sheet_title = trainer_course.course.name
            filename = f"{sanitize_title(trainer_course.course.name)}.xlsx"
        
        if build_sheet_for_course(wb, trainer_course.course, lectures, students, sheet_title):
            sheets_created += 1
    
    # Student-specific export: one student, optional course/batch/schedule filters
    elif report_type == 'student' and object_id:
        student = get_object_or_404(Student, id=object_id)
        course_id = request.GET.get('course') or request.GET.get('course_id')
        batch_id = request.GET.get('batch') or request.GET.get('batch_id')
        schedule = request.GET.get('schedule')
        # default schedule to the student's schedule if not provided
        schedule = schedule or getattr(student, 'schedule', None)
        
        if course_id:
            # Single course export for this student
            course = get_object_or_404(Course, id=course_id)
            # Only include lectures where this student has attendance - guarantees correct slot
            lectures = Lecture.objects.filter(
                attendances__student=student,
                trainer_course__course=course,
            )
            if batch_id:
                lectures = lectures.filter(trainer_course__batch_id=batch_id)
                batch_obj = Batch.objects.get(id=batch_id)
                sheet_title = f"{student.name} - {course.name} - {batch_obj.batch_number}"
                filename = f"{sanitize_title(student.name)} - {sanitize_title(course.name)} - {sanitize_title(batch_obj.batch_number)}.xlsx"
            else:
                sheet_title = f"{student.name} - {course.name}"
                filename = f"{sanitize_title(student.name)} - {sanitize_title(course.name)}.xlsx"
            # Single-student queryset
            students = Student.objects.filter(id=student.id)
            if build_sheet_for_course(wb, course, lectures, students, sheet_title):
                sheets_created += 1
        else:
            # Export all courses for this student
            courses = student.courses.all()
            for course in courses:
                # Get lectures for this specific course where student has attendance
                lectures = Lecture.objects.filter(
                    attendances__student=student,
                    trainer_course__course=course,
                )
                # Filter by student's batch if available
                if getattr(student, 'batch', None):
                    lectures = lectures.filter(
                        models.Q(trainer_course__batch__isnull=True) | 
                        models.Q(trainer_course__batch=student.batch)
                    )
                
                if lectures.exists():
                    batch_info = f" - Batch {student.batch.batch_number}" if getattr(student, 'batch', None) else ""
                    sheet_title = f"{student.name} - {course.name}{batch_info}"
                    # Single-student queryset for this course
                    students = Student.objects.filter(id=student.id)
                    if build_sheet_for_course(wb, course, lectures, students, sheet_title):
                        sheets_created += 1
            
            filename = f"{sanitize_title(student.name)} - All Courses.xlsx"

    # Trainer-wide export: one sheet per assignment course (respect batch on assignment)
    elif report_type == 'trainer' and ((is_trainer(request.user) and object_id is None) or (is_admin(request.user) and object_id is not None)):
        trainer = get_object_or_404(Trainer, id=object_id) if (object_id and is_admin(request.user)) else request.user.trainer_profile
        assignments = TrainerCourse.objects.filter(trainer=trainer, is_active=True).select_related('course', 'batch')
        for tc in assignments:
            # Only include lectures for this specific trainer assignment
            lectures = Lecture.objects.filter(trainer_course=tc).select_related('trainer_course__batch')
            schedule = getattr(tc, 'schedule', None)
            # Filter students by the trainer's specific assignment (batch and schedule)
            students = _students_for_assignment(tc.course, tc.batch, schedule)
            if tc.batch_id:
                title = f"{tc.course.name} - {tc.batch.batch_number}"
            else:
                title = f"{tc.course.name}"
            if build_sheet_for_course(wb, tc.course, lectures, students, title):
                sheets_created += 1
        filename = f"{sanitize_title(trainer.name)}.xlsx"

    # Batch export: sheets per course within the batch (only courses with attendance)
    elif report_type == 'batch' and object_id:
        batch = get_object_or_404(Batch, id=object_id)
        courses = Course.objects.filter(students__batch=batch).distinct()
        for c in courses:
            # Handle multiple assignments/slots (e.g., weekdays/weekend) for the same course within the batch
            tcs = list(TrainerCourse.objects.filter(course=c, batch=batch).select_related('trainer'))
            if not tcs:
                # Fallback: aggregate by course if no explicit assignment found
                lectures = Lecture.objects.filter(trainer_course__course=c, trainer_course__batch=batch)
                if not Attendance.objects.filter(lecture__in=lectures).exists():
                    continue
                students = _students_for_assignment(c, batch, None)
                title = f"{c.name}"
                if build_sheet_for_course(wb, c, lectures, students, title):
                    sheets_created += 1
                continue

            for tc in tcs:
                schedule = getattr(tc, 'schedule', None)
                # Only include lectures for this specific trainer assignment/slot
                lectures = Lecture.objects.filter(trainer_course=tc)
                if not Attendance.objects.filter(lecture__in=lectures).exists():
                    continue
                students = _students_for_assignment(c, batch, schedule)
                schedule_label = f" - {schedule.capitalize()}" if schedule else ""
                trainer_label = f" - {tc.trainer.name}" if getattr(tc, 'trainer', None) else ""
                title = f"{c.name}{schedule_label}{trainer_label}"
                if build_sheet_for_course(wb, c, lectures, students, title):
                    sheets_created += 1
        filename = f"Batch {sanitize_title(batch.batch_number)}.xlsx"

    elif is_admin(request.user):
        for c in Course.objects.all().order_by('name'):
            lectures = Lecture.objects.filter(trainer_course__course=c)
            if not Attendance.objects.filter(lecture__in=lectures).exists():
                continue
            students = c.students.select_related('batch').all()
            title = c.name
            if build_sheet_for_course(wb, c, lectures, students, title):
                sheets_created += 1
        filename = f"attendance_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        return JsonResponse({'success': False, 'message': 'Invalid report type'})

    # If no sheets were created, notify and redirect instead of erroring
    if sheets_created == 0:
        messages.info(request, 'No attendance found for the selected report.')
        if report_type == 'batch':
            return redirect('portal:batch_attendance_report')
        elif report_type == 'course':
            return redirect('portal:course_attendance_report')
        elif report_type == 'trainer':
            return redirect('portal:batch_attendance_report')
        return redirect('portal:admin_dashboard')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def student_details(request, student_id):
    student = get_object_or_404(Student.objects.select_related('batch'), id=student_id)
    # Enrolled courses
    courses = student.courses.all()
    # Build attendance summary per course with read-only history
    course_summaries = []
    for c in courses:
        # Lectures for this course, prefer matching student's batch if assignment had batch
        lectures_qs = Lecture.objects.filter(trainer_course__course=c)
        if getattr(student, 'batch', None):
            lectures_qs = lectures_qs.filter(models.Q(trainer_course__batch__isnull=True) | models.Q(trainer_course__batch=student.batch))
        attendances = Attendance.objects.filter(student=student, lecture__in=lectures_qs).select_related('lecture').order_by('lecture__date')
        total = attendances.count()
        present = attendances.filter(status='present').count()
        absent = attendances.filter(status='absent').count()
        # Build compact history list for template rendering (with 12-hour time)
        history = []
        for a in attendances:
            symbol = 'P' if a.status in ['present', 'late'] else ('A' if a.status == 'absent' else '-')
            start_str = a.lecture.start_time.strftime('%I:%M %p') if a.lecture.start_time else ''
            end_str = a.lecture.end_time.strftime('%I:%M %p') if a.lecture.end_time else ''
            history.append({
                'date': a.lecture.date,
                'status': a.status,
                'symbol': symbol,
                'start_time': start_str,
                'end_time': end_str,
            })
        course_summaries.append({
            'course': c,
            'total': total,
            'present': present,
            'absent': absent,
            'history': history,
        })
    context = {
        'student': student,
        'course_summaries': course_summaries,
    }
    return render(request, 'portal/admin/student_details.html', context)
